import os
import json
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import re
import shutil

# Для экспорта в PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Для экспорта в Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Для графика
import matplotlib.pyplot as plt

# Путь к шрифту
ASSETS_DIR = Path(__file__).parent / "assets"
CHAKRA_FONT_PATH = ASSETS_DIR / "ChakraPetch-Regular.ttf"
USE_CHAKRA_FONT = CHAKRA_FONT_PATH.exists()

if USE_CHAKRA_FONT:
    try:
        pdfmetrics.registerFont(TTFont('ChakraPetch', str(CHAKRA_FONT_PATH)))
    except Exception as e:
        messagebox.showwarning("Шрифт", f"Не удалось загрузить шрифт ChakraPetch:\n{e}")
        USE_CHAKRA_FONT = False

# Файл настроек (лежит рядом с программой)
SETTINGS_PATH = Path(__file__).parent / "settings.json"


def load_base_dir():
    """Загружает путь к базе из settings.json или запрашивает у пользователя при первом запуске."""
    if SETTINGS_PATH.exists():
        try:
            with open(SETTINGS_PATH, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                base_dir = Path(settings.get("base_dir", ""))
                if base_dir.is_dir():
                    return base_dir
        except Exception:
            pass

    # Файл не найден или путь недействителен → первый запуск
    root = tk.Tk()
    root.withdraw()  # Скрыть основное окно

    messagebox.showinfo("Первый запуск", "Пожалуйста, выберите папку с файлом base.json.")

    folder = filedialog.askdirectory(title="Выберите папку с базой данных (base.json)")
    root.destroy()

    if not folder:
        messagebox.showerror("Ошибка", "Папка не выбрана. Программа будет закрыта.")
        exit()

    base_dir = Path(folder)

    # Сохраняем настройки (игнорируем ошибку записи)
    try:
        save_base_dir(base_dir)
    except Exception:
        pass

    return base_dir


def save_base_dir(base_dir: Path):
    """Сохраняет путь к папке базы в settings.json."""
    try:
        with open(SETTINGS_PATH, 'w', encoding='utf-8') as f:
            json.dump({"base_dir": str(base_dir)}, f, ensure_ascii=False, indent=4)
    except Exception:
        # Игнорируем ошибку — нормально для readonly-пользователей
        pass


def ensure_base_exists(base_path: Path):
    """Создаёт папку и base.json, если не существует."""
    base_path.parent.mkdir(parents=True, exist_ok=True)
    if not base_path.exists():
        with open(base_path, 'w', encoding='utf-8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)


class RegistrumApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Registrum — Реестр счетов покупок")
        self.root.state('zoomed')

        # Загружаем путь к базе (запросит при первом запуске)
        self.base_dir = load_base_dir()
        self.base_path = self.base_dir / "base.json"

        # Определяем, есть ли права на запись
        self.readonly_mode = not self.can_write_to_base_dir()

        if self.readonly_mode:
            self.root.title("Registrum — Реестр счетов покупок [Только чтение]")
            messagebox.showinfo("Режим просмотра", "Обнаружен режим только для чтения.\n"
                                                   "Добавление, редактирование и удаление записей недоступны.")

        try:
            ensure_base_exists(self.base_path)
        except Exception as e:
            if not self.readonly_mode:
                messagebox.showerror("Ошибка", f"Не удалось создать базу данных:\n{e}")
                self.base_path = None
            # В readonly режиме ошибка создания игнорируется — файл должен уже существовать

        # === Поиск и кнопки на одном уровне ===
        top_frame = tk.Frame(root)
        top_frame.pack(pady=5, padx=20, fill=tk.X)

        tk.Label(top_frame, text="Поиск:", font=("Arial", 10)).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search_change)
        self.search_entry = tk.Entry(top_frame, textvariable=self.search_var, font=("Arial", 10), width=40)
        self.search_entry.pack(side=tk.LEFT, padx=(5, 10))

        tk.Button(top_frame, text="Очистить", command=self.clear_search).pack(side=tk.LEFT, padx=(0, 5))
        self.btn_backup = tk.Button(top_frame, text="Резерв", command=self.create_backup)
        self.btn_backup.pack(side=tk.LEFT, padx=(0, 20))

        # Кнопки управления (справа)
        btn_frame = tk.Frame(top_frame)
        btn_frame.pack(side=tk.RIGHT)

        self.btn_pdf = tk.Button(btn_frame, text="В PDF", command=self.export_to_pdf, width=12, height=1)
        self.btn_excel = tk.Button(btn_frame, text="В Excel", command=self.export_to_excel, width=12, height=1)
        self.btn_settings = tk.Button(btn_frame, text="Настройки", command=self.open_settings, width=12, height=1)
        self.btn_chart = tk.Button(btn_frame, text="График", command=self.show_chart, width=12, height=1)
        self.btn_info = tk.Button(btn_frame, text="Инфо", command=self.show_info, width=12, height=1)
        self.btn_exit = tk.Button(btn_frame, text="Выход", command=self.root.quit, width=12, height=1)

        for i, btn in enumerate([self.btn_pdf, self.btn_excel, self.btn_settings, self.btn_chart, self.btn_info, self.btn_exit]):
            btn.grid(row=0, column=i, padx=3)

        # Отключаем недоступные кнопки в readonly режиме
        if self.readonly_mode:
            self.btn_backup.config(state='disabled')
            self.btn_settings.config(state='disabled')

        # Таблица
        table_frame = tk.Frame(root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.columns = ["Дата", "Заказ", "Сумма", "Поставщик", "Плательщик", "Инициатор", "Обоснование", "Оплата", "Забрал", "Комментарии"]
        self.tree = ttk.Treeview(table_frame, columns=self.columns, show='headings')

        for col in self.columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            self.tree.column(col, width=100, anchor='center')

        v_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        h_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=v_scroll.set, xscroll=h_scroll.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Контекстное меню
        self.context_menu = tk.Menu(self.tree, tearoff=0)
        self.context_menu.add_command(label="Удалить", command=self.delete_selected)
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", self.on_double_click)

        # В readonly режиме удаляем пункт "Удалить"
        if self.readonly_mode:
            self.context_menu.delete(0)

        # Форма ввода — растянута на всю ширину
        form_frame = tk.Frame(root)
        form_frame.pack(pady=10, padx=20, fill=tk.X)

        tk.Label(form_frame, text="Форма добавления / редактирования заказа:", font=("Arial", 12, "bold")).pack(anchor='w')

        self.entries = {}
        grid_frame = tk.Frame(form_frame)
        grid_frame.pack(pady=5, fill=tk.X)

        short_fields = ["Дата", "Заказ", "Сумма", "Поставщик", "Плательщик", "Инициатор", "Оплата", "Забрал"]
        long_fields = ["Обоснование", "Комментарии"]

        for i, field in enumerate(short_fields):
            row = i // 2
            col = (i % 2) * 2
            tk.Label(grid_frame, text=f"{field}:", font=("Arial", 10)).grid(row=row, column=col, sticky='e', padx=5, pady=3)
            entry = tk.Entry(grid_frame, font=("Arial", 10), width=30)
            entry.grid(row=row, column=col+1, sticky='w', padx=5, pady=3)
            self.entries[field] = entry

        for field in long_fields:
            tk.Label(grid_frame, text=f"{field}:", font=("Arial", 10)).grid(
                row=len(short_fields)//2 + 1 + long_fields.index(field), column=0, sticky='ne', padx=5, pady=5)
            text = tk.Text(grid_frame, font=("Arial", 10), width=60, height=3)
            text.grid(row=len(short_fields)//2 + 1 + long_fields.index(field), column=1, columnspan=3, sticky='ew', padx=5, pady=5)
            self.entries[field] = text

        grid_frame.grid_columnconfigure(1, weight=1)
        grid_frame.grid_columnconfigure(3, weight=1)

        form_btn_frame = tk.Frame(form_frame)
        form_btn_frame.pack(pady=10)

        self.btn_new = tk.Button(form_btn_frame, text="Новый", command=self.clear_form, width=20, height=2)
        self.btn_save_order = tk.Button(form_btn_frame, text="Сохранить заказ", command=self.save_order, width=20, height=2)
        self.btn_cancel = tk.Button(form_btn_frame, text="Отмена", command=self.clear_form, width=20, height=2)

        self.btn_new.pack(side=tk.LEFT, padx=10)
        self.btn_save_order.pack(side=tk.LEFT, padx=10)
        self.btn_cancel.pack(side=tk.LEFT, padx=10)

        # Отключаем форму ввода в readonly режиме
        if self.readonly_mode:
            self.btn_new.config(state='disabled')
            self.btn_save_order.config(state='disabled')
            self.btn_cancel.config(state='disabled')
            for widget in self.entries.values():
                if isinstance(widget, tk.Entry):
                    widget.config(state='disabled')
                elif isinstance(widget, tk.Text):
                    widget.config(state='disabled')

        self.editing_index = None
        self.all_data = []
        self.load_table()
        self.clear_form()
        self.root.bind("<Escape>", lambda e: self.root.quit())

    def can_write_to_base_dir(self) -> bool:
        """Проверяет, можно ли писать в папку базы."""
        if not self.base_dir:
            return False
        try:
            test_file = self.base_dir / ".write_test_registrum"
            test_file.write_text("ok", encoding='utf-8')
            test_file.unlink()
            return True
        except (OSError, IOError, PermissionError):
            return False

    def load_table(self):
        self.all_data = self.load_data()
        self.refresh_table_view()
        self.auto_adjust_column_widths()

    def refresh_table_view(self, search_term=""):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if not search_term:
            filtered = self.all_data
        else:
            filtered = []
            for record in self.all_data:
                for col in self.columns:
                    if search_term.lower() in str(record.get(col, "")).lower():
                        filtered.append(record)
                        break

        for record in filtered:
            values = [record.get(col, "") for col in self.columns]
            self.tree.insert('', tk.END, values=values)

    def auto_adjust_column_widths(self):
        if not self.all_data:
            # Устанавливаем разумные значения по умолчанию, если данных нет
            default_widths = {
                "Дата": 80,
                "Заказ": 100,
                "Сумма": 90,
                "Поставщик": 130,
                "Плательщик": 120,
                "Инициатор": 90,
                "Оплата": 60,
                "Забрал": 60,
                "Обоснование": 400,
                "Комментарии": 250,
            }
            for col in self.columns:
                width = default_widths.get(col, 120)
                self.tree.column(col, width=width)
            return

        # Минимальные и максимальные ширины
        min_width = 60
        max_width = 300

        # Сначала определим базовую ширину по содержимому
        col_widths = {}
        for col in self.columns:
            header_len = len(col)
            max_content_len = max((len(str(record.get(col, ""))) for record in self.all_data), default=0)
            max_len = max(header_len, max_content_len)
            # Оценка ширины в пикселях (~8 px на символ)
            width_px = min(max(min_width, max_len * 8), max_width)
            col_widths[col] = width_px

        # Принудительно уменьшаем короткие столбцы
        fixed_short_width = 80
        for col in ["Дата", "Оплата", "Забрал"]:
            col_widths[col] = fixed_short_width

        # Увеличиваем "Обоснование" за счёт освобождённого места
        original_short_total = sum(min(col_widths[col], max_width) for col in ["Дата", "Оплата", "Забрал"])
        new_short_total = 3 * fixed_short_width
        saved_pixels = original_short_total - new_short_total

        # Добавим их к "Обоснованию"
        col_widths["Обоснование"] = min(max_width + saved_pixels, 500)  # максимум 500 px
        col_widths["Комментарии"] = min(col_widths.get("Комментарии", 250), 400)

        # Применяем ширины
        for col in self.columns:
            self.tree.column(col, width=col_widths[col])

    def on_search_change(self, *args):
        term = self.search_var.get()
        self.refresh_table_view(term)

    def clear_search(self):
        self.search_var.set("")
        self.refresh_table_view()

    def create_backup(self):
        if self.readonly_mode:
            messagebox.showwarning("Доступ запрещён", "Режим только для чтения. Создание резервной копии невозможно.")
            return

        if not self.base_path or not self.base_path.exists():
            messagebox.showwarning("Предупреждение", "Файл базы не существует!")
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_name = f"base_{timestamp}.json"
        backup_path = self.base_dir / backup_name

        try:
            shutil.copy2(self.base_path, backup_path)
            messagebox.showinfo("Успех", f"Резервная копия создана:\n{backup_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать резервную копию:\n{e}")

    def load_data(self):
        if not self.base_path or not self.base_path.exists():
            return []
        try:
            with open(self.base_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить базу:\n{e}")
            return []

    def save_data(self, data):
        if self.readonly_mode:
            return  # Ничего не делаем — или можно показать предупреждение, но уже проверено выше

        if not self.base_path:
            messagebox.showerror("Ошибка", "Путь к базе не задан.")
            return
        try:
            with open(self.base_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить базу:\n{e}")

    def clear_form(self):
        if self.readonly_mode:
            return

        self.editing_index = None
        for field in ["Дата", "Заказ", "Сумма", "Поставщик", "Плательщик", "Инициатор", "Оплата", "Забрал"]:
            self.entries[field].delete(0, tk.END)
            if field == "Дата":
                self.entries[field].insert(0, self.get_today())
            elif field == "Инициатор":
                self.entries[field].insert(0, "ИТ")
            elif field in ("Оплата", "Забрал"):
                self.entries[field].insert(0, "Да")
        for field in ["Обоснование", "Комментарии"]:
            self.entries[field].delete("1.0", tk.END)

    def get_today(self):
        return datetime.now().strftime("%d.%m.%Y")

    def on_double_click(self, event):
        selected = self.tree.selection()
        if not selected:
            return
        item = selected[0]
        index = self.tree.index(item)
        data = self.all_data
        if index >= len(data):
            return
        record = data[index]

        self.editing_index = index
        for field in ["Дата", "Заказ", "Сумма", "Поставщик", "Плательщик", "Инициатор", "Оплата", "Забрал"]:
            self.entries[field].delete(0, tk.END)
            self.entries[field].insert(0, record.get(field, ""))
        for field in ["Обоснование", "Комментарии"]:
            self.entries[field].delete("1.0", tk.END)
            self.entries[field].insert("1.0", record.get(field, ""))

    def save_order(self):
        if self.readonly_mode:
            messagebox.showwarning("Доступ запрещён", "Режим только для чтения. Изменения невозможны.")
            return

        record = {}
        for field in ["Дата", "Заказ", "Сумма", "Поставщик", "Плательщик", "Инициатор", "Оплата", "Забрал"]:
            record[field] = self.entries[field].get().strip()
        for field in ["Обоснование", "Комментарии"]:
            record[field] = self.entries[field].get("1.0", tk.END).strip()

        data = self.all_data

        if self.editing_index is not None:
            data[self.editing_index] = record
            self.save_data(data)
            self.load_table()
            messagebox.showinfo("Успех", "Запись успешно обновлена!")
        else:
            data.append(record)
            self.save_data(data)
            self.load_table()
            messagebox.showinfo("Успех", "Новый заказ успешно добавлен!")

        self.clear_form()

    def export_to_pdf(self):
        data = self.all_data
        if not data:
            messagebox.showwarning("Предупреждение", "База данных пуста!")
            return

        file_path = filedialog.asksaveasfilename(
            title="Сохранить как PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")]
        )
        if not file_path:
            return

        try:
            # Альбомная ориентация
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4), leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
            elements = []

            styles = getSampleStyleSheet()
            if USE_CHAKRA_FONT:
                title_style = styles['Title'].clone('CustomTitle')
                title_style.fontName = 'ChakraPetch'
                title_style.fontSize = 16
                title = Paragraph("Реестр счетов покупок (Registrum)", title_style)
            else:
                title = Paragraph("Реестр счетов покупок (Registrum)", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))

            table_data = [self.columns]
            for record in data:
                row = [str(record.get(col, "")) for col in self.columns]
                table_data.append(row)

            # Ширина страницы в альбомной ориентации
            total_width = landscape(A4)[0] - 72  # 72 = 36 + 36
            num_cols = len(self.columns)
            # Базовая ширина с учётом приоритета для "Обоснования"
            col_widths = []
            for col in self.columns:
                if col == "Обоснование":
                    col_widths.append(total_width * 0.30)  # 30% ширины страницы
                elif col in ("Комментарии", "Поставщик", "Плательщик"):
                    col_widths.append(total_width * 0.12)
                else:
                    col_widths.append(total_width * 0.07)

            # Нормализуем, чтобы сумма точно равнялась total_width
            actual_sum = sum(col_widths)
            if actual_sum > 0:
                col_widths = [w * total_width / actual_sum for w in col_widths]

            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'ChakraPetch' if USE_CHAKRA_FONT else 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]

            if USE_CHAKRA_FONT:
                table_style.append(('FONTNAME', (0, 1), (-1, -1), 'ChakraPetch'))

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle(table_style))

            elements.append(table)
            doc.build(elements)
            messagebox.showinfo("Успех", f"Файл PDF сохранён:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать PDF:\n{e}")

    def export_to_excel(self):
        data = self.all_data
        if not data:
            messagebox.showwarning("Предупреждение", "База данных пуста!")
            return

        file_path = filedialog.asksaveasfilename(
            title="Сохранить как Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр счетов"

            ws.append(self.columns)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for record in data:
                row = [record.get(col, "") for col in self.columns]
                ws.append(row)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            messagebox.showinfo("Успех", f"Файл Excel сохранён:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать Excel:\n{e}")

    def show_info(self):
        messagebox.showinfo("Информация", "Registrum v0.5 Автор Разин Г.В.")

    def open_settings(self):
        if self.readonly_mode:
            messagebox.showwarning("Доступ запрещён", "Режим только для чтения. Настройки недоступны.")
            return

        settings_win = tk.Toplevel(self.root)
        settings_win.title("Настройки")
        settings_win.geometry("600x150")
        settings_win.resizable(False, False)
        settings_win.grab_set()

        tk.Label(settings_win, text="Текущая папка базы данных:", font=("Arial", 10)).pack(pady=5)
        current_path = tk.Entry(settings_win, font=("Arial", 10), width=70)
        current_path.insert(0, str(self.base_dir))
        current_path.config(state='readonly')
        current_path.pack(pady=5)

        def change_path():
            new_dir = filedialog.askdirectory(title="Выберите папку для хранения базы данных (base.json)")
            if not new_dir:
                return
            new_dir_path = Path(new_dir)
            new_base_path = new_dir_path / "base.json"

            try:
                new_dir_path.mkdir(parents=True, exist_ok=True)
                test_file = new_dir_path / ".test_write"
                test_file.write_text("ok", encoding='utf-8')
                test_file.unlink()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Нет прав на запись в выбранную папку:\n{e}")
                return

            self.base_dir = new_dir_path
            self.base_path = new_base_path
            save_base_dir(self.base_dir)
            ensure_base_exists(self.base_path)
            self.load_table()
            self.clear_form()
            current_path.config(state='normal')
            current_path.delete(0, tk.END)
            current_path.insert(0, str(self.base_dir))
            current_path.config(state='readonly')
            messagebox.showinfo("Успех", f"Папка базы изменена на:\n{self.base_dir}")

        tk.Button(settings_win, text="Изменить путь", command=change_path, width=20).pack(pady=10)

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def delete_selected(self):
        if self.readonly_mode:
            messagebox.showwarning("Доступ запрещён", "Режим только для чтения. Удаление невозможно.")
            return

        selected = self.tree.selection()
        if not selected:
            return
        confirm = messagebox.askyesno("Подтверждение", "Вы уверены, что хотите удалить выбранную запись?")
        if not confirm:
            return

        index = self.tree.index(selected[0])
        self.all_data.pop(index)
        self.save_data(self.all_data)
        self.load_table()
        self.clear_form()
        messagebox.showinfo("Успех", "Запись удалена.")

    def sort_column(self, col):
        reverse = False
        if hasattr(self, '_last_sorted_col') and self._last_sorted_col == col:
            reverse = not getattr(self, '_last_sorted_reverse', False)
        else:
            reverse = False

        self._last_sorted_col = col
        self._last_sorted_reverse = reverse

        def _sort_key(record):
            val = record.get(col, "")
            if col == "Дата":
                # Парсим дату в (год, месяц, день) или возвращаем fallback
                if re.fullmatch(r'\d{2}\.\d{2}\.\d{4}', str(val)):
                    try:
                        d = datetime.strptime(val, "%d.%m.%Y")
                        return (d.year, d.month, d.day)
                    except:
                        pass
                return (9999, 99, 99)  # отправляем некорректные даты в конец
            else:
                # Для остальных столбцов — как раньше: число или строка
                try:
                    return float(str(val).replace(" ", "").replace(",", "."))
                except:
                    return str(val).lower()

        self.all_data.sort(key=_sort_key, reverse=reverse)
        self.refresh_table_view(self.search_var.get())

    def show_chart(self):
        data = self.all_data
        if not data:
            messagebox.showwarning("Предупреждение", "Нет данных для построения графика!")
            return

        yearly_totals = {}

        for record in data:
            date_str = record.get("Дата", "").strip()
            sum_str = record.get("Сумма", "").strip()

            year = None
            if re.match(r'\d{2}\.\d{2}\.\d{4}', date_str):
                try:
                    year = datetime.strptime(date_str, "%d.%m.%Y").year
                except:
                    pass

            if year is None:
                continue

            try:
                sum_val = float(sum_str.replace(" ", "").replace(",", "."))
            except:
                continue

            if year not in yearly_totals:
                yearly_totals[year] = 0.0
            yearly_totals[year] += sum_val

        if not yearly_totals:
            messagebox.showwarning("Предупреждение", "Не удалось извлечь год и сумму из данных!")
            return

        years = sorted(yearly_totals.keys())
        totals = [yearly_totals[y] for y in years]

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.bar(years, totals, color='steelblue')
        ax.set_title("Сравнение затрат по годам", fontsize=16)
        ax.set_xlabel("Год", fontsize=12)
        ax.set_ylabel("Сумма (руб.)", fontsize=12)
        ax.grid(axis='y', linestyle='--', alpha=0.7)

        # 🔑 Ключевое исправление: явно задаём метки по годам
        ax.set_xticks(years)
        ax.set_xticklabels([str(int(y)) for y in years])

        # Форматирование сумм на оси Y (с пробелами как разделитель тысяч)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x):,}'.replace(',', ' ')))

        # Подписи над столбцами
        for i, total in enumerate(totals):
            ax.text(years[i], total + max(totals) * 0.01, f'{int(total):,}'.replace(',', ' '),
                    ha='center', va='bottom', fontsize=9)

        manager = plt.get_current_fig_manager()
        try:
            manager.window.state('zoomed')
        except:
            try:
                manager.full_screen_toggle()
            except:
                pass

        plt.tight_layout()
        plt.show()


def main():
    root = tk.Tk()
    app = RegistrumApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()